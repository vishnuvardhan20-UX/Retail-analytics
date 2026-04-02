"""
generate_excel_report.py
Builds Retail_Analytics_Report.xlsx with 5 sheets:
  1. Dashboard (KPI summary)
  2. Monthly Revenue
  3. Top Products
  4. Region Analysis
  5. Category Breakdown
Uses openpyxl only — no external dependencies.
"""
import random
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ── reproduce same data as generate_data.py ───────────────────────────────────
random.seed(42)

START = date(2023, 1, 1)
END   = date(2024, 12, 31)

PRODUCTS = [
    (1,  "Wireless Earbuds",      "Electronics",  2499, 900),
    (2,  "Smartphone Cover",      "Electronics",   399, 100),
    (3,  "USB-C Hub",             "Electronics",  1899, 700),
    (4,  "Bluetooth Speaker",     "Electronics",  3499,1200),
    (5,  "Men's T-Shirt",         "Clothing",      599, 180),
    (6,  "Women's Kurta",         "Clothing",      899, 280),
    (7,  "Denim Jeans",           "Clothing",     1299, 450),
    (8,  "Running Shoes",         "Sports",       2799, 950),
    (9,  "Non-stick Pan",         "Home & Kitchen",999, 380),
    (10, "Steel Water Bottle",    "Home & Kitchen",449, 130),
    (11, "Air Fryer",             "Home & Kitchen",4999,2200),
    (12, "Fiction Novel",         "Books",         299,  80),
    (13, "Self-Help Book",        "Books",         349,  90),
    (14, "Yoga Mat",              "Sports",        799, 260),
    (15, "Cricket Bat",           "Sports",       1599, 600),
    (16, "Face Serum",            "Beauty",        899, 300),
    (17, "Moisturiser 50ml",      "Beauty",        499, 150),
    (18, "Laptop Stand",          "Electronics",  1299, 450),
    (19, "LED Desk Lamp",         "Home & Kitchen",799, 280),
    (20, "Protein Shaker Bottle", "Sports",        349, 100),
]
PROD_MAP  = {p[0]: p for p in PRODUCTS}
REGIONS   = {1:"North",2:"South",3:"East",4:"West",5:"Central"}
CITIES_BY = {1:["Delhi","Chandigarh","Jaipur","Lucknow"],
             2:["Chennai","Bangalore","Hyderabad","Kochi"],
             3:["Kolkata","Bhubaneswar","Patna","Guwahati"],
             4:["Mumbai","Pune","Ahmedabad","Surat"],
             5:["Bhopal","Nagpur","Raipur","Indore"]}
STATUSES  = ["Completed","Completed","Completed","Pending","Cancelled"]
PAYMENTS  = ["UPI","Credit Card","Debit Card","NetBanking","COD"]

def rand_date(s, e):
    return s + timedelta(days=random.randint(0,(e-s).days))

# customers
cust_region = {}
for cid in range(1,2001):
    rid = random.randint(1,5)
    cust_region[cid] = rid

# orders → aggregate
monthly_rev   = defaultdict(float)
product_units = defaultdict(int)
product_rev   = defaultdict(float)
region_rev    = defaultdict(float)
cat_rev       = defaultdict(float)
cat_units     = defaultdict(int)
payment_cnt   = defaultdict(int)
total_orders  = total_completed = total_cancelled = 0
total_revenue = 0.0
FNAMES=["Aarav","Aditi","Amit","Ananya","Arjun","Deepa","Gaurav","Isha",
        "Kiran","Manoj","Neha","Priya","Rahul","Riya","Rohan","Sanya",
        "Shruti","Suresh","Tanvi","Vikram"]
LNAMES=["Sharma","Verma","Patel","Singh","Kumar","Mehta","Joshi","Gupta",
        "Nair","Reddy","Iyer","Das","Shah","Rao","Mishra"]

for _ in range(1,2001):
    random.choice(FNAMES); random.choice(LNAMES)
    random.randint(1,5); rand_date(START,END)
    "".join([str(random.randint(0,9)) for _ in range(9)])

for oid in range(1, 8001):
    cid    = random.randint(1,2000)
    odate  = rand_date(START,END)
    status = random.choice(STATUSES)
    pay    = random.choice(PAYMENTS)
    total_orders += 1
    if status == "Completed": total_completed += 1
    if status == "Cancelled": total_cancelled += 1
    payment_cnt[pay] += 1
    n_items = random.randint(1,4)
    picked  = random.sample(PRODUCTS, n_items)
    for prod in picked:
        pid,pname,cat,up,cp = prod
        qty  = random.randint(1,5)
        disc = random.choice([0,0,0,5,10,15])
        rev  = qty * up * (1 - disc/100)
        if status == "Completed":
            month_key = odate.strftime("%Y-%m")
            monthly_rev[month_key]   += rev
            product_units[pid]       += qty
            product_rev[pid]         += rev
            region_rev[cust_region[cid]] += rev
            cat_rev[cat]             += rev
            cat_units[cat]           += qty
            total_revenue            += rev
        random.random()  # return

# ── sorted data ───────────────────────────────────────────────────────────────
months_sorted  = sorted(monthly_rev.keys())
month_revenues = [round(monthly_rev[m], 2) for m in months_sorted]

top_products = sorted(product_rev.items(), key=lambda x: x[1], reverse=True)[:10]

region_data = [(REGIONS[rid], round(rev,2))
               for rid,rev in sorted(region_rev.items(), key=lambda x:-x[1])]

cat_data = [(cat, round(cat_rev[cat],2), cat_units[cat])
            for cat in sorted(cat_rev, key=lambda x:-cat_rev[x])]

avg_order_val = round(total_revenue / max(total_completed,1), 2)

# ── styling helpers ───────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT     = "E8F4FD"
WHITE      = "FFFFFF"
GRAY       = "F2F2F2"
GREEN      = "375623"
GREEN_FILL = "E2EFDA"

def hdr_cell(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def data_cell(ws, row, col, val, fmt=None, bold=False, bg=None, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=10, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt: c.number_format = fmt
    if bg:  c.fill = PatternFill("solid", start_color=bg)
    return c

def kpi_box(ws, row, col, label, value, fmt="#,##0"):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col,
                   end_row=row+1, end_column=col+1)
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(bold=True, color=WHITE, size=10, name="Arial")
    lc.fill = PatternFill("solid", start_color=MID_BLUE)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws.cell(row=row+1, column=col, value=value)
    vc.font = Font(bold=True, color=DARK_BLUE, size=14, name="Arial")
    vc.fill = PatternFill("solid", start_color=LIGHT_BLUE)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = fmt
    return vc

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(r,c).border = border

# ── workbook ─────────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — Dashboard
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Dashboard"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 2
for col in ["B","C","D","E","F","G","H","I","J","K"]:
    ws1.column_dimensions[col].width = 16
ws1.row_dimensions[1].height = 10
ws1.row_dimensions[2].height = 40
ws1.row_dimensions[3].height = 10

# Title
ws1.merge_cells("B2:K2")
tc = ws1["B2"]
tc.value = "🛒  Retail Analytics — BI Dashboard (2023–2024)"
tc.font = Font(bold=True, size=18, color=WHITE, name="Arial")
tc.fill = PatternFill("solid", start_color=DARK_BLUE)
tc.alignment = Alignment(horizontal="center", vertical="center")

# KPI row
for r in [4,5,6]: ws1.row_dimensions[r].height = 28
kpi_box(ws1, 4, 2,  "Total Revenue (₹)",   total_revenue,          "#,##0")
kpi_box(ws1, 4, 4,  "Completed Orders",     total_completed,        "#,##0")
kpi_box(ws1, 4, 6,  "Avg Order Value (₹)",  avg_order_val,          "#,##0.00")
kpi_box(ws1, 4, 8,  "Unique Customers",     2000,                   "#,##0")
kpi_box(ws1, 4, 10, "Total Products",       20,                     "#,##0")

# Mini monthly table for chart
ws1.cell(8,2).value = "Month"
ws1.cell(8,3).value = "Revenue"
ws1.cell(8,2).font = ws1.cell(8,3).font = Font(bold=True, name="Arial")
for i, (m, r) in enumerate(zip(months_sorted, month_revenues)):
    ws1.cell(9+i, 2).value = m
    ws1.cell(9+i, 3).value = r
    ws1.cell(9+i, 3).number_format = "#,##0"

# Revenue Line chart
line = LineChart()
line.title   = "Monthly Revenue Trend"
line.style   = 10
line.y_axis.title = "Revenue (₹)"
line.x_axis.title = "Month"
line.width = 18; line.height = 11
data_ref = Reference(ws1, min_col=3, min_row=8,
                     max_row=8+len(months_sorted))
cats_ref = Reference(ws1, min_col=2, min_row=9,
                     max_row=8+len(months_sorted))
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.series[0].graphicalProperties.line.solidFill = MID_BLUE
line.series[0].graphicalProperties.line.width = 20000
ws1.add_chart(line, "E7")

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — Monthly Revenue
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Revenue")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2
widths = [14,18,18,18]
for i,w in enumerate(widths,2):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("B1:E1")
t2 = ws2["B1"]
t2.value = "Monthly Revenue Report"
t2.font  = Font(bold=True, size=14, color=WHITE, name="Arial")
t2.fill  = PatternFill("solid", start_color=DARK_BLUE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

headers = ["Month","Revenue (₹)","MoM Growth %","Running Total (₹)"]
for col,h in enumerate(headers,2):
    hdr_cell(ws2, 2, col, h, bg=MID_BLUE)
ws2.row_dimensions[2].height = 24

running = 0
for i,(m,r) in enumerate(zip(months_sorted, month_revenues)):
    row = 3+i
    running += r
    prev = month_revenues[i-1] if i > 0 else None
    growth = round((r - prev)/prev*100, 2) if prev else None
    bg = ACCENT if i%2==0 else WHITE
    data_cell(ws2, row, 2, m,       align="center", bg=bg)
    data_cell(ws2, row, 3, r,       fmt="#,##0.00", align="right", bg=bg)
    data_cell(ws2, row, 4, growth,  fmt='0.00"%"',  align="right", bg=bg)
    data_cell(ws2, row, 5, running, fmt="#,##0.00", align="right", bg=bg)
    ws2.row_dimensions[row].height = 18

apply_border(ws2, 2, 2+len(months_sorted), 2, 5)

# Bar chart
bar = BarChart()
bar.type  = "col"; bar.style = 10
bar.title = "Monthly Revenue (₹)"
bar.y_axis.title = "Revenue"
bar.x_axis.title = "Month"
bar.width = 20; bar.height = 12
dr = Reference(ws2, min_col=3, min_row=2, max_row=2+len(months_sorted))
cr = Reference(ws2, min_col=2, min_row=3, max_row=2+len(months_sorted))
bar.add_data(dr, titles_from_data=True)
bar.set_categories(cr)
bar.series[0].graphicalProperties.solidFill = MID_BLUE
ws2.add_chart(bar, f"B{4+len(months_sorted)}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3 — Top Products
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Top Products")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 2
for col,w in zip(["B","C","D","E","F"],[5,28,18,14,14]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells("B1:F1")
t3 = ws3["B1"]
t3.value = "Top 10 Products by Revenue"
t3.font  = Font(bold=True, size=14, color=WHITE, name="Arial")
t3.fill  = PatternFill("solid", start_color=DARK_BLUE)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

for col,h in zip([2,3,4,5,6],["Rank","Product","Category","Units Sold","Revenue (₹)"]):
    hdr_cell(ws3, 2, col, h, bg=MID_BLUE)

for i,(pid,rev) in enumerate(top_products):
    _,pname,cat,up,cp = PROD_MAP[pid]
    row = 3+i
    bg = GREEN_FILL if i < 3 else (ACCENT if i%2==0 else WHITE)
    data_cell(ws3, row, 2, i+1,             align="center", bg=bg, bold=(i<3))
    data_cell(ws3, row, 3, pname,           align="left",   bg=bg, bold=(i<3))
    data_cell(ws3, row, 4, cat,             align="left",   bg=bg)
    data_cell(ws3, row, 5, product_units[pid], fmt="#,##0", align="right", bg=bg)
    data_cell(ws3, row, 6, round(rev,2),    fmt="#,##0.00", align="right", bg=bg, bold=(i<3))
    ws3.row_dimensions[row].height = 20

apply_border(ws3, 2, 2+len(top_products), 2, 6)

# Horizontal bar for top products
hbar = BarChart()
hbar.type  = "bar"; hbar.style = 10
hbar.title = "Top 10 Products — Revenue"
hbar.y_axis.title = "Product"
hbar.x_axis.title = "Revenue (₹)"
hbar.width = 22; hbar.height = 14
dr3 = Reference(ws3, min_col=6, min_row=2, max_row=12)
cr3 = Reference(ws3, min_col=3, min_row=3, max_row=12)
hbar.add_data(dr3, titles_from_data=True)
hbar.set_categories(cr3)
hbar.series[0].graphicalProperties.solidFill = "2E75B6"
ws3.add_chart(hbar, "B14")

# ═══════════════════════════════════════════════════════════════
# SHEET 4 — Region Analysis
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Region Analysis")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 2
for col,w in zip(["B","C","D"],[18,18,18]):
    ws4.column_dimensions[col].width = w

ws4.merge_cells("B1:D1")
t4 = ws4["B1"]
t4.value = "Revenue by Region"
t4.font  = Font(bold=True, size=14, color=WHITE, name="Arial")
t4.fill  = PatternFill("solid", start_color=DARK_BLUE)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 32

for col,h in zip([2,3,4],["Region","Revenue (₹)","Share %"]):
    hdr_cell(ws4, 2, col, h, bg=MID_BLUE)

total_r = sum(r for _,r in region_data)
for i,(reg,rev) in enumerate(region_data):
    row = 3+i
    bg = ACCENT if i%2==0 else WHITE
    data_cell(ws4, row, 2, reg,                   align="left",  bg=bg)
    data_cell(ws4, row, 3, rev,   fmt="#,##0.00", align="right", bg=bg)
    data_cell(ws4, row, 4, round(rev/total_r*100,2), fmt='0.00"%"', align="right", bg=bg)
    ws4.row_dimensions[row].height = 20

apply_border(ws4, 2, 2+len(region_data), 2, 4)

# Pie chart
pie = PieChart()
pie.title  = "Revenue Share by Region"
pie.style  = 10
pie.width  = 16; pie.height = 12
dr4 = Reference(ws4, min_col=3, min_row=2, max_row=2+len(region_data))
cr4 = Reference(ws4, min_col=2, min_row=3, max_row=2+len(region_data))
pie.add_data(dr4, titles_from_data=True)
pie.set_categories(cr4)
ws4.add_chart(pie, "F2")

# ═══════════════════════════════════════════════════════════════
# SHEET 5 — Category Breakdown
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Category Breakdown")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 2
for col,w in zip(["B","C","D","E"],[22,18,14,16]):
    ws5.column_dimensions[col].width = w

ws5.merge_cells("B1:E1")
t5 = ws5["B1"]
t5.value = "Revenue & Units by Category"
t5.font  = Font(bold=True, size=14, color=WHITE, name="Arial")
t5.fill  = PatternFill("solid", start_color=DARK_BLUE)
t5.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 32

for col,h in zip([2,3,4,5],["Category","Revenue (₹)","Units Sold","Revenue Share %"]):
    hdr_cell(ws5, 2, col, h, bg=MID_BLUE)

total_cat = sum(r for _,r,_ in cat_data)
for i,(cat,rev,units) in enumerate(cat_data):
    row = 3+i
    bg = ACCENT if i%2==0 else WHITE
    data_cell(ws5, row, 2, cat,                     align="left",  bg=bg)
    data_cell(ws5, row, 3, rev,   fmt="#,##0.00",   align="right", bg=bg)
    data_cell(ws5, row, 4, units, fmt="#,##0",       align="right", bg=bg)
    data_cell(ws5, row, 5, round(rev/total_cat*100,2), fmt='0.00"%"', align="right", bg=bg)
    ws5.row_dimensions[row].height = 20

apply_border(ws5, 2, 2+len(cat_data), 2, 5)

# Bar chart for categories
cbar = BarChart()
cbar.type  = "col"; cbar.style = 10
cbar.title = "Revenue by Category"
cbar.y_axis.title = "Revenue (₹)"
cbar.width = 18; cbar.height = 12
dr5 = Reference(ws5, min_col=3, min_row=2, max_row=2+len(cat_data))
cr5 = Reference(ws5, min_col=2, min_row=3, max_row=2+len(cat_data))
cbar.add_data(dr5, titles_from_data=True)
cbar.set_categories(cr5)
cbar.series[0].graphicalProperties.solidFill = "375623"
ws5.add_chart(cbar, "B10")

# ── save ─────────────────────────────────────────────────────────────────────
wb.save("Retail_Analytics_Report.xlsx")
print("✅ Retail_Analytics_Report.xlsx created")
print(f"   Sheets: {', '.join(wb.sheetnames)}")
print(f"   Total Revenue : ₹{total_revenue:,.2f}")
print(f"   Completed     : {total_completed:,} orders")
print(f"   Avg Order Val : ₹{avg_order_val:,.2f}")
